# bot.py
import logging
from pathlib import Path
from telegram import Update
from telegram.ext import ApplicationBuilder, ContextTypes, CommandHandler, MessageHandler, filters
from PIL import Image
import pytesseract
import fitz  # PyMuPDF
from docx import Document
import openpyxl
from pdf2image import convert_from_path

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Константы (должны быть определены где-то выше в коде)
DOWNLOAD_DIR = Path("downloads")
DOWNLOAD_DIR.mkdir(exist_ok=True)

# Функция send_to_gigachat должна быть определена
def send_to_gigachat(prompt: str) -> str:
    # Реализация отправки в GigaChat
    pass

# Функция start должна быть определена
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Бот запущен!")

# Функция handle_text должна быть определена
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Обрабатываю текстовое сообщение...")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка фотографий"""
    photo = update.message.photo[-1]  # Берем самую большую версию фото
    file = await context.bot.get_file(photo.file_id)
    local_path = DOWNLOAD_DIR / f"photo_{photo.file_id}.jpg"
    
    try:
        await file.download_to_drive(custom_path=str(local_path))
        text = ocr_image(local_path)
        prompt = f"Пользователь прислал изображение. Распознанный текст:\n\n{text}\n\nВопрос: {update.message.caption or 'Нет явного вопроса — прокомментируй содержимое.'}"
        reply = send_to_gigachat(prompt)
        await update.message.reply_text(reply)
    except Exception as e:
        logger.exception('OCR failed')
        await update.message.reply_text(f"Ошибка при OCR: {e}")
    finally:
        # Удаляем временный файл
        if local_path.exists():
            local_path.unlink()

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка документов"""
    doc = update.message.document
    file = await context.bot.get_file(doc.file_id)
    local_path = DOWNLOAD_DIR / f"{doc.file_id}_{doc.file_name}"
    
    try:
        await file.download_to_drive(custom_path=str(local_path))
        
        if doc.mime_type == 'application/pdf' or (doc.file_name and doc.file_name.lower().endswith('.pdf')):
            await update.message.reply_text("PDF получен — извлекаю текст...")
            try:
                text = extract_text_from_pdf(local_path)
                prompt = f"Пользователь прислал PDF. Извлечённый текст:\n\n{text[:4000]}\n\nВопрос: {update.message.caption or 'Прокомментируй содержание.'}"
                reply = send_to_gigachat(prompt)
                await update.message.reply_text(reply)
            except Exception as e:
                logger.exception('PDF processing failed')
                await update.message.reply_text(f"Ошибка при обработке PDF: {e}")

        elif doc.mime_type in [
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword'] or (doc.file_name and doc.file_name.lower().endswith(('.doc', '.docx'))):
            await update.message.reply_text("DOCX получен — извлекаю текст...")
            try:
                text = extract_text_from_docx(local_path)
                prompt = f"Пользователь прислал DOCX. Извлечённый текст:\n\n{text[:4000]}\n\nВопрос: {update.message.caption or 'Прокомментируй содержание.'}"
                reply = send_to_gigachat(prompt)
                await update.message.reply_text(reply)
            except Exception as e:
                logger.exception('DOCX processing failed')
                await update.message.reply_text(f"Ошибка при обработке DOCX: {e}")

        elif doc.mime_type in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'] or (doc.file_name and doc.file_name.lower().endswith(('.xls', '.xlsx'))):
            await update.message.reply_text("XLSX получен — извлекаю данные...")
            try:
                text = extract_text_from_xlsx(local_path)
                prompt = f"Пользователь прислал XLSX. Извлечённые данные:\n\n{text[:4000]}\n\nВопрос: {update.message.caption or 'Прокомментируй данные.'}"
                reply = send_to_gigachat(prompt)
                await update.message.reply_text(reply)
            except Exception as e:
                logger.exception('XLSX processing failed')
                await update.message.reply_text(f"Ошибка при обработке XLSX: {e}")
        else:
            await update.message.reply_text("Поддерживаются только PDF, DOCX и XLSX.")
    
    except Exception as e:
        logger.exception('Document download failed')
        await update.message.reply_text(f"Ошибка при загрузке файла: {e}")
    finally:
        # Удаляем временный файл
        if local_path.exists():
            local_path.unlink()

def ocr_image(path: Path) -> str:
    """OCR для изображений"""
    try:
        img = Image.open(path)
        if img.mode != 'RGB':
            img = img.convert('RGB')
        text = pytesseract.image_to_string(img, lang='rus+eng')  # Добавлена поддержка русского
        return text.strip() if text else "Текст не распознан"
    except Exception as e:
        logger.error(f"OCR error: {e}")
        return f"Ошибка OCR: {e}"

def extract_text_from_pdf(path: Path) -> str:
    """Извлечение текста из PDF"""
    try:
        doc = fitz.open(str(path))
        full_text = []
        need_ocr_pages = []
        
        for page_num in range(doc.page_count):
            page = doc.load_page(page_num)
            text = page.get_text().strip()
            if text:
                full_text.append(f"Страница {page_num + 1}:\n{text}")
            else:
                need_ocr_pages.append(page_num)
        
        # OCR для страниц без текста
        if need_ocr_pages:
            try:
                images = convert_from_path(str(path), first_page=min(need_ocr_pages)+1, last_page=max(need_ocr_pages)+1)
                for i, page_num in enumerate(need_ocr_pages):
                    img = images[i]
                    txt = pytesseract.image_to_string(img, lang='rus+eng')
                    full_text.append(f"Страница {page_num + 1} (OCR):\n{txt}")
            except Exception as ocr_error:
                logger.error(f"PDF OCR failed: {ocr_error}")
                full_text.append(f"Страницы {need_ocr_pages}: OCR не удался")
        
        doc.close()
        return "\n\n".join(full_text) if full_text else "Текст не найден в PDF"
    
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        return f"Ошибка извлечения PDF: {e}"

def extract_text_from_docx(path: Path) -> str:
    """Извлечение текста из DOCX"""
    try:
        doc = Document(str(path))
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return text if text else "Текст не найден в DOCX"
    except Exception as e:
        logger.error(f"DOCX extraction error: {e}")
        return f"Ошибка извлечения DOCX: {e}"

def extract_text_from_xlsx(path: Path) -> str:
    """Извлечение текста из XLSX"""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        all_text = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else '' for cell in row])
                if row_text.strip():  # Пропускаем пустые строки
                    rows.append(row_text)
            
            if rows:  # Добавляем только непустые листы
                all_text.append(f"[Лист: {sheet}]\n" + "\n".join(rows))
        
        wb.close()
        return "\n\n---ЛИСТ---\n\n".join(all_text) if all_text else "Данные не найдены в XLSX"
    
    except Exception as e:
        logger.error(f"XLSX extraction error: {e}")
        return f"Ошибка извлечения XLSX: {e}"

def main():
    """Основная функция запуска бота"""
    # Получаем токены из переменных окружения
    TELEGRAM_TOKEN = "YOUR_TELEGRAM_TOKEN"  # Должно быть из os.getenv()
    GIGA_CHAT_TOKEN = "YOUR_GIGACHAT_TOKEN"  # Должно быть из os.getenv()
    
    if not TELEGRAM_TOKEN or not GIGA_CHAT_TOKEN:
        logger.error('TELEGRAM_TOKEN or GIGA_CHAT_TOKEN not set in environment')
        print('Set TELEGRAM_TOKEN and GIGA_CHAT_TOKEN in env or .env file (see .env.example)')
        return

    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    # Регистрируем обработчики
    app.add_handler(CommandHandler('start', start))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    logger.info('Starting bot...')
    app.run_polling()

if __name__ == '__main__':
    main()
